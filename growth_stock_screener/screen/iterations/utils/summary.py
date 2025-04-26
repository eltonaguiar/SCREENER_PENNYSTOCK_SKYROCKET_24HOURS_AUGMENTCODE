import os
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from .outfiles import open_outfile

def create_summary_file():
    """
    Create a summary Excel file with tabs for each stage of the screening process.
    """
    # Define the stages and their corresponding JSON files
    stages = [
        {"name": "1. Relative Strength", "file": "relative_strengths"},
        {"name": "2. Liquidity", "file": "liquidity"},
        {"name": "3. Trend", "file": "trend"},
        {"name": "4. Revenue Growth", "file": "revenue_growth"},
        {"name": "5. Final Results", "file": "institutional_accumulation"}
    ]
    
    # Create a new workbook
    wb = Workbook()
    
    # Remove the default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), 
        right=Side(style="thin"), 
        top=Side(style="thin"), 
        bottom=Side(style="thin")
    )
    
    # Create a sheet for each stage
    for stage in stages:
        try:
            # Try to open the JSON file for this stage
            df = open_outfile(stage["file"])
            
            # Create a new sheet
            ws = wb.create_sheet(title=stage["name"])
            
            # Write the data to the sheet
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    
                    # Apply styles to header row
                    if r_idx == 1:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
                        cell.border = thin_border
                    else:
                        cell.border = thin_border
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Freeze the header row
            ws.freeze_panes = "A2"
            
        except Exception as e:
            print(f"Error creating sheet for {stage['name']}: {e}")
    
    # Create a summary sheet
    summary_sheet = wb.create_sheet(title="Summary", index=0)
    
    # Add summary information
    summary_sheet["A1"] = "Growth Stock Screener Summary"
    summary_sheet["A1"].font = Font(bold=True, size=16)
    summary_sheet.merge_cells("A1:D1")
    
    summary_sheet["A3"] = "Date:"
    summary_sheet["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    summary_sheet["A5"] = "Stage"
    summary_sheet["B5"] = "Stocks Remaining"
    summary_sheet["C5"] = "% of Total"
    
    # Apply header styles
    for cell in summary_sheet["5:5"]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Get the counts for each stage
    row = 6
    total_stocks = 0
    
    for i, stage in enumerate(stages):
        try:
            df = open_outfile(stage["file"])
            count = len(df)
            
            # Get the total from the first stage
            if i == 0:
                total_stocks = count
            
            summary_sheet[f"A{row}"] = stage["name"]
            summary_sheet[f"B{row}"] = count
            
            # Calculate percentage of total
            if total_stocks > 0:
                percentage = (count / total_stocks) * 100
                summary_sheet[f"C{row}"] = f"{percentage:.2f}%"
            else:
                summary_sheet[f"C{row}"] = "N/A"
            
            # Apply border
            for cell in summary_sheet[f"{row}:{row}"]:
                cell.border = thin_border
            
            row += 1
        except Exception as e:
            print(f"Error adding summary for {stage['name']}: {e}")
    
    # Auto-adjust column widths
    for column in summary_sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = (max_length + 2) * 1.2
        summary_sheet.column_dimensions[column_letter].width = adjusted_width
    
    # Save the workbook
    time_string = datetime.now().strftime("%Y-%m-%d %H-%M-%S")
    file_path = os.path.join(os.getcwd(), f"screen_summary {time_string}.xlsx")
    wb.save(file_path)
    
    return file_path
